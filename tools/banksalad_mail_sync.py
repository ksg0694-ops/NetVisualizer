#!/usr/bin/env python3
"""Import BankSalad Gmail export attachments into NetVisualizer.

The worker intentionally uses read-only Gmail access. Raw mail bodies, ZIP
attachments, decrypted workbooks, and ZIP passwords are never persisted.
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import io
import json
import logging
import os
import re
import sys
import unicodedata
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from email.header import decode_header
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

import pyzipper
import requests
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import load_workbook


LOGGER = logging.getLogger("banksalad-mail-sync")
SOURCE = "banksalad_gmail"
DEFAULT_MAIL_QUERY = (
    "from:export-noreply@banksalad.com has:attachment filename:zip newer_than:30d"
)
GMAIL_READONLY_SCOPE = "https://www.googleapis.com/auth/gmail.readonly"
MAX_ZIP_BYTES = 30 * 1024 * 1024
MAX_WORKBOOK_BYTES = 50 * 1024 * 1024


HEADER_ALIASES = {
    "날짜": "date",
    "일자": "date",
    "거래일": "date",
    "거래일자": "date",
    "date": "date",
    "시간": "time",
    "거래시간": "time",
    "time": "time",
    "타입": "type",
    "유형": "type",
    "구분": "type",
    "type": "type",
    "대분류": "category",
    "분류": "category",
    "카테고리": "category",
    "category": "category",
    "소분류": "subcategory",
    "상세분류": "subcategory",
    "subcategory": "subcategory",
    "내용": "memo",
    "내역": "memo",
    "메모": "memo",
    "적요": "memo",
    "description": "memo",
    "memo": "memo",
    "금액": "amount",
    "거래금액": "amount",
    "amount": "amount",
    "출금": "withdrawal",
    "출금액": "withdrawal",
    "withdrawal": "withdrawal",
    "debit": "withdrawal",
    "입금": "deposit",
    "입금액": "deposit",
    "deposit": "deposit",
    "credit": "deposit",
    "결제금액": "expense_amount",
    "이용금액": "expense_amount",
    "사용금액": "expense_amount",
    "paymentamount": "expense_amount",
    "통화": "currency",
    "currency": "currency",
    "결제수단": "method",
    "계좌": "method",
    "계좌명": "method",
    "카드": "method",
    "카드명": "method",
    "method": "method",
    "account": "method",
}
EXPECTED_HEADERS = set(HEADER_ALIASES.values())


class SyncError(RuntimeError):
    """Base class for sanitized operational failures."""


class AttachmentNotFoundError(SyncError):
    pass


class WorkbookFormatError(SyncError):
    pass


class ZipPasswordError(SyncError):
    pass


@dataclass(frozen=True)
class NormalizedTransaction:
    date: str
    time: str | None
    type: str
    category: str
    subcategory: str
    memo: str
    amount: int
    currency: str
    method: str

    def fingerprint(self) -> str:
        # Category and subcategory describe a transaction; they do not identify
        # the underlying bank event. Keeping them out of the digest prevents a
        # later classification correction from inserting a second transaction.
        payload = [
            self.date,
            self.time or "",
            self.type,
            self.amount,
            normalize_dedupe_text(self.memo),
            normalize_dedupe_text(self.method),
            self.currency,
        ]
        serialized = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        return hashlib.sha256(serialized.encode("utf-8")).hexdigest()

    def to_supabase_row(self, owner_user_id: str, gmail_message_id: str) -> dict[str, Any]:
        row = asdict(self)
        row.update(
            {
                "user_id": owner_user_id,
                "source": SOURCE,
                "source_message_id": gmail_message_id,
                "dedupe_key": self.fingerprint(),
            }
        )
        return row


@dataclass(frozen=True)
class ParseResult:
    sheet_name: str
    rows_seen: int
    invalid_rows: int
    duplicate_rows: int
    transactions: list[NormalizedTransaction]


@dataclass(frozen=True)
class GmailAttachment:
    message_id: str
    internal_date_ms: int
    filename: str
    data: bytes

    @property
    def name_hash(self) -> str:
        return hashlib.sha256(self.filename.encode("utf-8")).hexdigest()


def require_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise SyncError(f"missing_environment:{name}")
    return value


def normalize_text(value: Any, fallback: str = "") -> str:
    text = unicodedata.normalize("NFKC", str(value if value is not None else ""))
    text = re.sub(r"\s+", " ", text).strip()
    return text or fallback


def normalize_dedupe_text(value: Any) -> str:
    text = normalize_text(value).lower()
    return re.sub(r"[\s_\-.,/\\()[\]{}:;'\"`~!@#$%^&*+=|<>?]", "", text)


def normalize_header(value: Any) -> str:
    key = normalize_text(value).lower()
    key = re.sub(r"[\s_\-./()[\]{}]", "", key)
    return HEADER_ALIASES.get(key, key)


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    raw = normalize_text(value)
    if not raw:
        return ""
    raw = raw.replace("년", "-").replace("월", "-").replace("일", "")
    raw = re.sub(r"\s+", "", raw)
    match = re.match(r"^(\d{4})(?:[-./]?)(\d{1,2})(?:[-./]?)(\d{1,2})", raw)
    if not match:
        return ""
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3))).isoformat()
    except ValueError:
        return ""


def normalize_time(value: Any) -> str | None:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.replace(microsecond=0).isoformat()
    if isinstance(value, (float, Decimal)) and 0 <= float(value) < 1:
        seconds = round(float(value) * 24 * 60 * 60) % (24 * 60 * 60)
        return f"{seconds // 3600:02d}:{(seconds % 3600) // 60:02d}:{seconds % 60:02d}"

    raw = normalize_text(value)
    if not raw:
        return None
    match = re.search(r"(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?", raw)
    if match:
        hour, minute, second = (int(match.group(1)), int(match.group(2)), int(match.group(3) or 0))
    else:
        digits = re.sub(r"\D", "", raw)
        if len(digits) not in (3, 4, 6):
            return None
        digits = digits.zfill(4 if len(digits) == 3 else len(digits))
        hour, minute = int(digits[:2]), int(digits[2:4])
        second = int(digits[4:6]) if len(digits) == 6 else 0
    if hour > 23 or minute > 59 or second > 59:
        return None
    return f"{hour:02d}:{minute:02d}:{second:02d}"


def parse_number(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return round(float(value))
    cleaned = normalize_text(value)
    cleaned = cleaned.replace(",", "").replace("원", "")
    cleaned = re.sub(r"[^\d.\-]", "", cleaned)
    if cleaned in ("", "-", ".", "-."):
        return None
    try:
        return round(Decimal(cleaned))
    except (InvalidOperation, ValueError):
        return None


def infer_type(raw_type: Any, amount: int, method: str) -> str:
    key = re.sub(r"\s+", "", normalize_text(raw_type).lower())
    if key in {"수입", "입금", "income", "deposit", "credit"}:
        return "수입"
    if key in {"지출", "출금", "결제", "expense", "spend", "withdrawal", "debit"}:
        return "지출"
    if key in {"이체", "transfer"}:
        return "이체"
    if "카드" in method and amount > 0:
        return "지출"
    return "지출" if amount < 0 else "수입"


def normalize_amount(values: Mapping[str, Any], tx_type: str) -> int | None:
    amount = parse_number(values.get("amount"))
    withdrawal = parse_number(values.get("withdrawal"))
    deposit = parse_number(values.get("deposit"))
    expense = parse_number(values.get("expense_amount"))
    if amount in (None, 0):
        if withdrawal not in (None, 0):
            amount = -abs(withdrawal)
        elif deposit not in (None, 0):
            amount = abs(deposit)
        elif expense not in (None, 0):
            amount = -abs(expense)
    if amount in (None, 0):
        return None
    if tx_type == "지출":
        return -abs(amount)
    if tx_type == "수입":
        return abs(amount)
    return round(amount)


def row_to_transaction(headers: Sequence[str], row: Sequence[Any]) -> NormalizedTransaction | None:
    values: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header or index >= len(row):
            continue
        value = row[index]
        if value not in (None, "") and header not in values:
            values[header] = value

    tx_date = normalize_date(values.get("date"))
    method = normalize_text(values.get("method"))
    provisional_amount = parse_number(values.get("amount")) or 0
    tx_type = infer_type(values.get("type"), provisional_amount, method)
    amount = normalize_amount(values, tx_type)
    if not tx_date or amount in (None, 0):
        return None

    return NormalizedTransaction(
        date=tx_date,
        time=normalize_time(values.get("time")),
        type=tx_type,
        category=normalize_text(values.get("category"), "미분류"),
        subcategory=normalize_text(values.get("subcategory"), "미분류"),
        memo=normalize_text(values.get("memo"), "미분류 거래"),
        amount=amount,
        currency=normalize_text(values.get("currency"), "KRW").upper(),
        method=method,
    )


def header_score(row: Sequence[Any]) -> int:
    return sum(1 for value in row if normalize_header(value) in EXPECTED_HEADERS)


def choose_ledger_sheet(workbook: Any) -> tuple[Any, int]:
    candidates: list[tuple[int, int, Any, int]] = []
    for sheet_index, worksheet in enumerate(workbook.worksheets):
        best_score = 0
        best_header_index = -1
        for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=30, values_only=True)):
            score = header_score(row)
            if score > best_score:
                best_score = score
                best_header_index = row_index
        name_bonus = 3 if re.search(r"가계부|거래|내역", worksheet.title, re.I) else 0
        second_sheet_bonus = 2 if sheet_index == 1 else 0
        total_score = best_score * 10 + name_bonus + second_sheet_bonus
        candidates.append((total_score, best_score, worksheet, best_header_index))

    candidates.sort(key=lambda item: item[0], reverse=True)
    if not candidates or candidates[0][1] < 3 or candidates[0][3] < 0:
        raise WorkbookFormatError("transaction_header_not_found")
    _, _, worksheet, zero_based_header_index = candidates[0]
    return worksheet, zero_based_header_index + 1


def parse_workbook_bytes(workbook_bytes: bytes) -> ParseResult:
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    try:
        worksheet, header_row_number = choose_ledger_sheet(workbook)
        raw_headers = next(
            worksheet.iter_rows(
                min_row=header_row_number,
                max_row=header_row_number,
                values_only=True,
            )
        )
        headers = [normalize_header(value) for value in raw_headers]
        rows_seen = 0
        invalid_rows = 0
        duplicate_rows = 0
        transactions: list[NormalizedTransaction] = []
        fingerprints: set[str] = set()

        for row in worksheet.iter_rows(min_row=header_row_number + 1, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue
            rows_seen += 1
            transaction = row_to_transaction(headers, row)
            if transaction is None:
                invalid_rows += 1
                continue
            fingerprint = transaction.fingerprint()
            if fingerprint in fingerprints:
                duplicate_rows += 1
                continue
            fingerprints.add(fingerprint)
            transactions.append(transaction)

        if not transactions:
            raise WorkbookFormatError("no_valid_transactions")
        return ParseResult(
            sheet_name=worksheet.title,
            rows_seen=rows_seen,
            invalid_rows=invalid_rows,
            duplicate_rows=duplicate_rows,
            transactions=transactions,
        )
    finally:
        workbook.close()


def decrypt_workbook_from_zip(zip_bytes: bytes, password: str) -> bytes:
    if len(zip_bytes) > MAX_ZIP_BYTES:
        raise SyncError("zip_too_large")
    try:
        with pyzipper.AESZipFile(io.BytesIO(zip_bytes)) as archive:
            archive.setpassword(password.encode("utf-8"))
            candidates = [
                info
                for info in archive.infolist()
                if not info.is_dir()
                and info.filename.lower().endswith(".xlsx")
                and not info.filename.startswith("__MACOSX/")
                and not Path(info.filename).name.startswith("~$")
            ]
            if not candidates:
                raise WorkbookFormatError("xlsx_not_found")
            candidates.sort(key=lambda info: info.file_size, reverse=True)
            selected = candidates[0]
            if selected.file_size > MAX_WORKBOOK_BYTES:
                raise SyncError("workbook_too_large")
            workbook_bytes = archive.read(selected)
    except RuntimeError as error:
        raise ZipPasswordError("zip_decryption_failed") from error
    except pyzipper.BadZipFile as error:
        raise WorkbookFormatError("invalid_zip") from error
    return workbook_bytes


def decode_mime_header(value: str | None) -> str:
    if not value:
        return ""
    parts: list[str] = []
    for chunk, encoding in decode_header(value):
        if isinstance(chunk, bytes):
            parts.append(chunk.decode(encoding or "utf-8", errors="replace"))
        else:
            parts.append(chunk)
    return "".join(parts)


def iter_message_parts(payload: Mapping[str, Any]) -> Iterator[Mapping[str, Any]]:
    yield payload
    for part in payload.get("parts") or []:
        yield from iter_message_parts(part)


def decode_gmail_data(encoded: str) -> bytes:
    padding = "=" * (-len(encoded) % 4)
    return base64.urlsafe_b64decode(encoded + padding)


def build_gmail_service() -> Any:
    credentials = Credentials(
        token=None,
        refresh_token=require_env("GMAIL_OAUTH_REFRESH_TOKEN"),
        token_uri="https://oauth2.googleapis.com/token",
        client_id=require_env("GMAIL_OAUTH_CLIENT_ID"),
        client_secret=require_env("GMAIL_OAUTH_CLIENT_SECRET"),
        scopes=[GMAIL_READONLY_SCOPE],
    )
    return build("gmail", "v1", credentials=credentials, cache_discovery=False)


def list_matching_message_ids(service: Any, query: str, max_messages: int) -> list[str]:
    message_ids: list[str] = []
    page_token: str | None = None
    while len(message_ids) < max_messages:
        response = (
            service.users()
            .messages()
            .list(
                userId="me",
                q=query,
                maxResults=min(100, max_messages - len(message_ids)),
                pageToken=page_token,
            )
            .execute()
        )
        message_ids.extend(item["id"] for item in response.get("messages", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            break
    return message_ids[:max_messages]


def fetch_zip_attachment(service: Any, message_id: str) -> GmailAttachment:
    message = (
        service.users().messages().get(userId="me", id=message_id, format="full").execute()
    )
    payload = message.get("payload") or {}
    headers = {item.get("name", "").lower(): item.get("value", "") for item in payload.get("headers", [])}
    sender = decode_mime_header(headers.get("from")).lower()
    if "export-noreply@banksalad.com" not in sender:
        raise AttachmentNotFoundError("unexpected_sender")

    for part in iter_message_parts(payload):
        filename = decode_mime_header(part.get("filename"))
        if not filename.lower().endswith(".zip"):
            continue
        body = part.get("body") or {}
        data = body.get("data")
        attachment_id = body.get("attachmentId")
        if attachment_id:
            attachment = (
                service.users()
                .messages()
                .attachments()
                .get(userId="me", messageId=message_id, id=attachment_id)
                .execute()
            )
            data = attachment.get("data")
        if not data:
            continue
        return GmailAttachment(
            message_id=message_id,
            internal_date_ms=int(message.get("internalDate") or 0),
            filename=filename,
            data=decode_gmail_data(data),
        )
    raise AttachmentNotFoundError("zip_attachment_not_found")


class SupabaseRest:
    def __init__(self, url: str, service_role_key: str, owner_user_id: str) -> None:
        self.url = url.rstrip("/")
        self.owner_user_id = owner_user_id
        self.session = requests.Session()
        self.session.headers.update(
            {
                "apikey": service_role_key,
                "Authorization": f"Bearer {service_role_key}",
                "Content-Type": "application/json",
            }
        )

    def _request(self, method: str, path: str, **kwargs: Any) -> requests.Response:
        response = self.session.request(method, f"{self.url}/rest/v1/{path}", timeout=45, **kwargs)
        if not response.ok:
            raise SyncError(f"supabase_http_{response.status_code}")
        return response

    def completed_message_ids(self, message_ids: Sequence[str]) -> set[str]:
        if not message_ids:
            return set()
        quoted = ",".join(f'"{message_id}"' for message_id in message_ids)
        response = self._request(
            "GET",
            "banksalad_sync_runs",
            params={
                "select": "gmail_message_id",
                "user_id": f"eq.{self.owner_user_id}",
                "status": "eq.completed",
                "gmail_message_id": f"in.({quoted})",
            },
        )
        return {row["gmail_message_id"] for row in response.json()}

    def existing_fingerprints(self) -> set[str]:
        fingerprints: set[str] = set()
        start = 0
        page_size = 1000
        select = "date,time,type,category,subcategory,memo,amount,currency,method,dedupe_key"
        while True:
            response = self._request(
                "GET",
                "transactions",
                params={
                    "select": select,
                    "user_id": f"eq.{self.owner_user_id}",
                    "order": "date.asc",
                },
                headers={"Range": f"{start}-{start + page_size - 1}"},
            )
            rows = response.json()
            for row in rows:
                transaction = NormalizedTransaction(
                    date=normalize_date(row.get("date")),
                    time=normalize_time(row.get("time")),
                    type=normalize_text(row.get("type")),
                    category=normalize_text(row.get("category"), "미분류"),
                    subcategory=normalize_text(row.get("subcategory"), "미분류"),
                    memo=normalize_text(row.get("memo"), "미분류 거래"),
                    amount=int(Decimal(str(row.get("amount") or 0))),
                    currency=normalize_text(row.get("currency"), "KRW").upper(),
                    method=normalize_text(row.get("method")),
                )
                # Always recompute with the current identity contract. Legacy
                # stored digests included category fields and cannot safely be
                # compared with the classification-agnostic digest.
                fingerprints.add(transaction.fingerprint())
            if len(rows) < page_size:
                break
            start += page_size
        return fingerprints

    def insert_transactions(self, rows: Sequence[Mapping[str, Any]]) -> int:
        inserted = 0
        for start in range(0, len(rows), 500):
            batch = list(rows[start : start + 500])
            response = self._request(
                "POST",
                "transactions",
                params={"on_conflict": "user_id,source,dedupe_key"},
                headers={"Prefer": "resolution=ignore-duplicates,return=representation"},
                data=json.dumps(batch, ensure_ascii=False),
            )
            inserted += len(response.json())
        return inserted

    def upsert_run(self, payload: Mapping[str, Any]) -> None:
        row = {"user_id": self.owner_user_id, **payload}
        self._request(
            "POST",
            "banksalad_sync_runs",
            params={"on_conflict": "user_id,gmail_message_id"},
            headers={"Prefer": "resolution=merge-duplicates,return=minimal"},
            data=json.dumps(row, ensure_ascii=False),
        )


def sanitize_error_code(error: BaseException) -> str:
    if isinstance(error, SyncError) and str(error):
        return re.sub(r"[^a-zA-Z0-9_:-]", "_", str(error))[:120]
    return error.__class__.__name__[:120]


def build_database_from_env() -> SupabaseRest:
    return SupabaseRest(
        require_env("SUPABASE_URL"),
        require_env("SUPABASE_SERVICE_ROLE_KEY"),
        require_env("NETVISUALIZER_OWNER_USER_ID"),
    )


def process_attachment(
    attachment: GmailAttachment,
    password: str,
    database: SupabaseRest | None,
    existing_fingerprints: set[str],
    dry_run: bool,
) -> dict[str, Any]:
    workbook_bytes = decrypt_workbook_from_zip(attachment.data, password)
    result = parse_workbook_bytes(workbook_bytes)
    new_transactions = [
        transaction
        for transaction in result.transactions
        if transaction.fingerprint() not in existing_fingerprints
    ]
    skipped_existing = len(result.transactions) - len(new_transactions)

    inserted = 0
    if not dry_run:
        if database is None:
            raise SyncError("database_required")
        rows = [
            transaction.to_supabase_row(database.owner_user_id, attachment.message_id)
            for transaction in new_transactions
        ]
        inserted = database.insert_transactions(rows) if rows else 0
        existing_fingerprints.update(transaction.fingerprint() for transaction in new_transactions)

    return {
        "sheet_name": result.sheet_name,
        "rows_seen": result.rows_seen,
        "rows_valid": len(result.transactions),
        "rows_inserted": inserted,
        "rows_duplicate_file": result.duplicate_rows,
        "rows_duplicate_existing": skipped_existing,
        "rows_invalid": result.invalid_rows,
    }


def run_gmail_sync(args: argparse.Namespace) -> int:
    password = require_env("BANKSALAD_ZIP_PASSWORD")
    service = build_gmail_service()
    database = None if args.dry_run else build_database_from_env()
    message_ids = list_matching_message_ids(service, args.mail_query, args.max_messages)
    if not message_ids:
        LOGGER.info("No matching BankSalad messages found.")
        return 0

    completed = set() if database is None else database.completed_message_ids(message_ids)
    attachments = [
        fetch_zip_attachment(service, message_id)
        for message_id in message_ids
        if message_id not in completed
    ]
    attachments.sort(key=lambda item: item.internal_date_ms)
    if not attachments:
        LOGGER.info("All matching BankSalad messages were already completed.")
        return 0

    existing = set() if database is None else database.existing_fingerprints()
    failures = 0
    for attachment in attachments:
        base_run = {
            "gmail_message_id": attachment.message_id,
            "attachment_name_hash": attachment.name_hash,
            "status": "processing",
            "started_at": datetime.now().astimezone().isoformat(),
            "finished_at": None,
            "error_code": None,
        }
        if database is not None:
            database.upsert_run(base_run)
        try:
            summary = process_attachment(
                attachment,
                password,
                database,
                existing,
                args.dry_run,
            )
            LOGGER.info("Message %s summary: %s", attachment.message_id, json.dumps(summary, ensure_ascii=False))
            if database is not None:
                database.upsert_run(
                    {
                        **base_run,
                        **summary,
                        "status": "completed",
                        "finished_at": datetime.now().astimezone().isoformat(),
                    }
                )
        except Exception as error:  # audit, then fail the job after remaining messages
            failures += 1
            error_code = sanitize_error_code(error)
            LOGGER.error("Message %s failed: %s", attachment.message_id, error_code)
            if database is not None:
                database.upsert_run(
                    {
                        **base_run,
                        "status": "failed",
                        "finished_at": datetime.now().astimezone().isoformat(),
                        "error_code": error_code,
                    }
                )
    return 1 if failures else 0


def run_local_zip(args: argparse.Namespace) -> int:
    password = require_env("BANKSALAD_ZIP_PASSWORD")
    zip_bytes = Path(args.zip_file).read_bytes()
    workbook_bytes = decrypt_workbook_from_zip(zip_bytes, password)
    result = parse_workbook_bytes(workbook_bytes)
    summary = {
        "sheet_name": result.sheet_name,
        "rows_seen": result.rows_seen,
        "rows_valid": len(result.transactions),
        "rows_duplicate_file": result.duplicate_rows,
        "rows_invalid": result.invalid_rows,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Read and parse Gmail attachments without DB writes")
    parser.add_argument("--mail-query", default=DEFAULT_MAIL_QUERY)
    parser.add_argument("--max-messages", type=int, default=20)
    parser.add_argument("--zip-file", help="Parse one local encrypted ZIP without Gmail or DB access")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    args = parse_args(argv or sys.argv[1:])
    if args.max_messages < 1 or args.max_messages > 100:
        raise SyncError("max_messages_out_of_range")
    if args.zip_file:
        return run_local_zip(args)
    return run_gmail_sync(args)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        LOGGER.error("Sync failed: %s", sanitize_error_code(error))
        raise SystemExit(1) from None
